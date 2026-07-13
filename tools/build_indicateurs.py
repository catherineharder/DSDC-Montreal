#!/usr/bin/env python3
"""Génère assets/data/indicateurs.data.js — indicateurs agrégés selon trois
découpages : tables de quartier (tq), arrondissements et villes liées (vdm),
réseaux locaux de services (sq).

Indicateurs :
1. defavo — indice de défavorisation matérielle et sociale (IDMS, INSPQ),
   recensements 2011, 2016, 2021. Quintiles régionaux (RSS 06 — Montréal).
2. iemv — indice d'équité des milieux de vie (Ville de Montréal, version 2026),
   nombre de vulnérabilités cumulées (0-6) par aire de diffusion.

Affectation des aires de diffusion (AD) aux territoires :
- tq  : point représentatif de l'AD dans les polygones WGS84 des 32 tables
        (tables_de_quartier_32_2024.geojson). AD hors territoire exclues.
- vdm : nom d'arrondissement / ville liée porté par chaque AD dans le GeoJSON
        IEMV (2021) ; pour 2011 et 2016, AD rattachée au territoire de l'AD 2021
        la plus proche (approximation aux frontières, négligeable à cette échelle).
- sq  : code de RLS porté par chaque AD dans les tables INSPQ (exact, toutes
        années). Pour 2011, l'AD scindée est rattachée au RLS dominant.

L'IDMS est une mesure relative (quintiles recalculés à chaque recensement) ;
l'IEMV n'est pas comparable d'une version à l'autre (pas de tendance).

Usage : python3 tools/build_indicateurs.py
"""

import io
import json
import math
import re
import sys
import unicodedata
import zipfile
from pathlib import Path

import csv

import openpyxl
import shapefile  # pyshp
from pyproj import Transformer
from shapely.geometry import Point, Polygon, shape
from shapely.prepared import prep
from shapely.strtree import STRtree

ROOT = Path(__file__).resolve().parent.parent          # github-pages/
DRSP = ROOT.parent                                     # dossier DRSP
IND = DRSP / "Indicators"
TDQ_GEOJSON = DRSP / "tables_de_quartier_32_2024.geojson"
SILHOUETTE_GEOJSON = DRSP / "Concertations et Partenariats" / "montreal-silhouette.geojson"
CAPITAL_XLSX = IND / "données catherine 13.07.26.xlsx"
VILLE_DATA_JS = ROOT / "assets" / "data" / "ville-montreal.data.js"
OUT = ROOT / "assets" / "data" / "indicateurs.data.js"
OUT_GEO = ROOT / "assets" / "data" / "indicateurs-iemv-geo.data.js"

# dimensions de l'IEMV : champ ACP_* du GeoJSON -> clé courte + libellé
IEMV_DIMS = [
    ("cult", "ACP_CultSportLoisir", "Culture, sports et loisirs"),
    ("prox", "ACP_proximite", "Ressources et proximité"),
    ("secu", "ACP_securite", "Sécurité urbaine"),
    ("envi", "ACP_envi", "Environnemental"),
    ("eco", "ACP_eco", "Économique"),
    ("soci", "ACP_soci", "Social"),
]

INSPQ_XLSX = {
    2011: IND / "A-DonnéesIDMS_Qc2011_fr" / "1.a TableCorrespondances_Qc2011_NouveauDecoupage_fr.xlsx",
    2016: IND / "A-DonnéesIDMS_Qc2016_fr" / "1. TableCorrespondancesCompleteQuebec2016.xlsx",
    2021: IND / "A-DonnéesIDMS_Qc2021_fr" / "1. TableCorrespondancesCompleteQuebec2021_fr.xlsx",
}
POPCOL = {2011: "ADPOP2011", 2016: "ADPOP2016", 2021: "ADPOP2021"}
FRACCOL = {2011: "FRAC_RLS", 2016: "FRAC_CLSC", 2021: "FRAC_CLSC"}

AD_GEOJSON_2021 = [IND / "resultats_iemv_v2026_agglo.geojson",
                   IND / "resultats_iemv_v2026_agglo.json"]
AD_SHAPEZIP = {2016: IND / "lda_000b16a_e.zip", 2011: IND / "gda_000b11a_e.zip"}
STATCAN_EPSG = 3347  # Lambert conique conforme (fichiers des limites StatCan)


def slugify(s):
    s = s.replace("–", "-").replace("—", "-")  # tirets demi-cadratin (noms Ville)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()


def extract_js_object(text, marker):
    """Extrait l'objet JSON qui suit `marker` dans un fichier .data.js."""
    i = text.index(marker) + len(marker)
    while text[i] != "{":
        i += 1
    depth = 0
    for j in range(i, len(text)):
        if text[j] == "{":
            depth += 1
        elif text[j] == "}":
            depth -= 1
            if depth == 0:
                return json.loads(text[i:j + 1])
    raise ValueError(marker)


# ---------------------------------------------------------------- INSPQ ----
def load_inspq(year):
    """AD (RSS 06) -> {pop, qm, qs, rls}. AD scindée : rangée dominante (FRAC max)."""
    path = INSPQ_XLSX[year]
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Données"]
    rows = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(rows))}
    best_frac = {}
    out = {}
    for r in rows:
        if str(r[hdr["RSS"]]) != "06":
            continue
        ad = str(r[hdr["AD"]])
        frac = r[hdr[FRACCOL[year]]] or 0
        if ad in out and frac <= best_frac[ad]:
            continue
        qm, qs = r[hdr["QuintMatRSS"]], r[hdr["QuintSocRSS"]]
        if qm is None or qs is None:
            continue
        best_frac[ad] = frac
        out[ad] = {"pop": r[hdr[POPCOL[year]]] or 0, "qm": int(qm), "qs": int(qs),
                   "rls": str(r[hdr["RLS"]])}
    return out


def load_rls_2021():
    """AD -> RLS pour toutes les AD 2021 (sans filtre sur les quintiles)."""
    wb = openpyxl.load_workbook(INSPQ_XLSX[2021], read_only=True)
    ws = wb["Données"]
    rows = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(rows))}
    out = {}
    for r in rows:
        if str(r[hdr["RSS"]]) == "06":
            out.setdefault(str(r[hdr["AD"]]), str(r[hdr["RLS"]]))
    return out


# ------------------------------------------------------------- géométrie ----
def load_iemv():
    """GeoJSON IEMV : AD 2021 -> (Point WGS84, nom arrondissement/ville, score, pop)."""
    path = next((p for p in AD_GEOJSON_2021 if p.exists()), None)
    if path is None:
        sys.exit("ERREUR : GeoJSON IEMV introuvable dans DRSP/Indicators/.")
    gj = json.load(open(path, encoding="utf-8"))
    crs_name = ((gj.get("crs") or {}).get("properties") or {}).get("name", "")
    m = re.search(r"EPSG::?(\d+)", crs_name)
    epsg = int(m.group(1)) if m else 4326
    tr = (Transformer.from_crs(f"EPSG:{epsg}", "EPSG:4326", always_xy=True).transform
          if epsg != 4326 else None)
    out = {}
    for f in gj["features"]:
        p = f["properties"]
        ad = str(p.get("ADIDU")).split(".")[0]
        if ad in out:
            continue
        geom = shape(f["geometry"])
        pt = geom.representative_point()
        if tr:
            pt = Point(*tr(pt.x, pt.y))
        score = p.get("Indice_emv")
        score = None if score is None or (isinstance(score, float) and math.isnan(score)) else int(score)
        mpc = p.get("MPC21")
        mpc = None if mpc is None or (isinstance(mpc, float) and math.isnan(mpc)) else float(mpc)
        acp = {}
        for key, field, _lbl in IEMV_DIMS:
            v = p.get(field)
            acp[key] = None if v is None or (isinstance(v, float) and math.isnan(v)) else int(v)
        out[ad] = {"pt": pt, "nom": p.get("NOM"), "score": score, "mpc": mpc,
                   "pop": int(p.get("tot_pop") or 0), "acp": acp,
                   "geom": geom, "tr": tr}
    print(f"IEMV : {path.name}, {len(out)} AD")
    return out


def points_statcan(year):
    """ad -> Point WGS84 (île de Montréal), zip StatCan ou dossier décompressé."""
    zpath = AD_SHAPEZIP[year]
    folder = zpath.with_suffix("")
    if zpath.exists():
        z = zipfile.ZipFile(zpath)
        base = next(n[:-4] for n in z.namelist() if n.lower().endswith(".shp"))
        rdr = shapefile.Reader(
            shp=io.BytesIO(z.read(base + ".shp")),
            dbf=io.BytesIO(z.read(base + ".dbf")),
            shx=io.BytesIO(z.read(base + ".shx")), encoding="latin-1")
    elif folder.is_dir():
        shp = next(folder.glob("*.shp"), None)
        if shp is None:
            return None
        rdr = shapefile.Reader(str(shp), encoding="latin-1")
    else:
        return None
    fields = [f[0] for f in rdr.fields[1:]]
    i_dauid = next(i for i, f in enumerate(fields) if f.upper() == "DAUID")
    proj = abs(rdr.bbox[0]) > 360 or abs(rdr.bbox[1]) > 360
    tr = (Transformer.from_crs(f"EPSG:{STATCAN_EPSG}", "EPSG:4326", always_xy=True).transform
          if proj else None)
    pts = {}
    for sr in rdr.iterShapeRecords():
        dauid = str(sr.record[i_dauid])
        if not dauid.startswith("2466") or dauid in pts:
            continue
        p = shape(sr.shape.__geo_interface__).representative_point()
        pts[dauid] = Point(*tr(p.x, p.y)) if tr else p
    print(f"géométrie {year} : {len(pts)} AD (île de Montréal)")
    return pts


# ----------------------------------------------------------- affectations ----
def tq_assigner():
    tdq = json.load(open(TDQ_GEOJSON, encoding="utf-8"))
    tables = [(slugify(f["properties"]["nom"]), prep(shape(f["geometry"])))
              for f in tdq["features"]]

    def assign(pt):
        for slug, pg in tables:
            if pg.contains(pt):
                return slug
        return None
    return assign


def vdm_name_to_slug():
    """slugify(nom IEMV) -> slug de la carte Ville (arrondissements + villes liées)."""
    text = VILLE_DATA_JS.read_text(encoding="utf-8")
    boroughs = extract_js_object(text, "const BOROUGHS =")
    geometry_suburbs = extract_js_object(text, "suburbs:")
    m = {}
    for slug, rec in boroughs.items():
        m[slugify(rec["name"])] = slug
    for slug in geometry_suburbs:
        m[slugify(slug)] = slug
    return m


# ---------------------------------------------- IEMV : mosaïque d'AD (SVG) ----
W, H, PAD = 1000, 875, 2.0
LAT0 = 45.55


def make_projector(bounds, lat0=LAT0):
    """Même projection que build_tables_quartier.py (calibrée sur la silhouette)."""
    minlon, minlat, maxlon, maxlat = bounds
    k = math.cos(math.radians(lat0))
    x0, x1 = minlon * k, maxlon * k
    cw, ch = (x1 - x0), (maxlat - minlat)
    aw, ah = W - 2 * PAD, H - 2 * PAD
    scale = min(aw / cw, ah / ch)
    ox = PAD + (aw - scale * cw) / 2
    oy = PAD + (ah - scale * ch) / 2

    def project(lon, lat):
        return (ox + (lon * k - x0) * scale, oy + (maxlat - lat) * scale)
    return project


def _poly_parts(geom):
    t = geom.geom_type
    if t == "Polygon":
        return [geom]
    if t in ("MultiPolygon", "GeometryCollection"):
        out = []
        for g in geom.geoms:
            out += _poly_parts(g)
        return out
    return []


def geom_to_path(geom, nd=1):
    """Géométrie shapely (espace SVG) -> attribut d du path."""
    fmt = f"%.{nd}f"
    parts = []
    for p in _poly_parts(geom):
        for ring in [p.exterior, *p.interiors]:
            pts = list(ring.coords)
            parts.append("M" + " ".join(
                (fmt % x).rstrip("0").rstrip(".") + "," + (fmt % y).rstrip("0").rstrip(".")
                for x, y in pts[:-1]) + "Z")
    return "".join(parts)


def iemv_class(v, score):
    """Classe d'une AD pour une dimension (v = flag ACP) ou l'ensemble (v=None).
       pv = vulnérable et prioritaire, vnp = vulnérable non prioritaire,
       nv = non vulnérable."""
    if score is None:
        return None
    if v is None:  # ensemble : niveau selon le nombre de vulnérabilités
        return "pv" if score >= 4 else ("vnp" if score == 3 else "nv")
    if v == 1:
        return "pv" if score >= 4 else "vnp"
    return "nv"


def build_iemv_geo(iemv):
    """Dissout les AD par option (ensemble + 6 dimensions) et par classe,
       projetées dans le viewBox 1000x875 du site."""
    from shapely.ops import unary_union
    sil = json.load(open(SILHOUETTE_GEOJSON, encoding="utf-8"))
    sil_geom = unary_union([make_valid_safe(shape(f["geometry"]))
                            for f in sil["features"]])
    project = make_projector(sil_geom.bounds)

    # projeter chaque AD une seule fois
    projected = {}
    for ad, v in iemv.items():
        if v["score"] is None:
            continue
        tr = v["tr"]
        polys = []
        for p in _poly_parts(v["geom"]):
            ext = ([tr(x, y) for x, y in p.exterior.coords] if tr
                   else list(p.exterior.coords))
            ext = [project(lon, lat) for lon, lat in ext]
            try:
                polys.append(make_valid_safe(Polygon(ext)).buffer(0))
            except Exception:
                continue
        if polys:
            projected[ad] = unary_union(polys)

    options = [("ens", None)] + [(key, key) for key, _f, _l in IEMV_DIMS]
    out = {}
    for opt, dim in options:
        groups = {"pv": [], "vnp": [], "nv": []}
        for ad, g in projected.items():
            v = iemv[ad]
            cls = iemv_class(None if dim is None else v["acp"].get(dim), v["score"])
            if cls:
                groups[cls].append(g)
        out[opt] = {}
        for cls, geoms in groups.items():
            if not geoms:
                continue
            u = unary_union(geoms).simplify(0.4)
            out[opt][cls] = geom_to_path(u)
        print(f"iemv geo {opt} : " + " | ".join(f"{c} {len(g)} AD" for c, g in groups.items()))
    return out


def make_valid_safe(g):
    from shapely.validation import make_valid
    try:
        return make_valid(g)
    except Exception:
        return g.buffer(0)


# --------------------------------------- Capital social (xlsx Infocentre) ----
def load_capital():
    """Extrait sentiment d'appartenance (ESCC), satisfaction vie sociale et
       degré de solitude (EQSP) du classeur Infocentre fourni."""
    if not CAPITAL_XLSX.exists():
        print("capital : classeur absent — section sautée")
        return None
    wb = openpyxl.load_workbook(CAPITAL_XLSX, data_only=True)
    num = lambda v: isinstance(v, (int, float))

    # sentiment d'appartenance fort (très + plutôt), Montréal vs Québec, par cycle
    ws = wb["senti-appartenance"]
    cycle, terr, series = None, None, {}
    for row in ws.iter_rows(values_only=True):
        c0 = str(row[0]) if row[0] else ""
        m = re.search(r"ESCC (\d{4})-(\d{4})", c0)
        if m:
            cycle = m.group(1) + "-" + m.group(2)[2:]
            terr = None
            continue
        if "Montréal" in c0:
            terr = "mtl"
        elif "Québec" in c0:
            terr = "qc"
        if cycle and terr and row[1] == "Total" and row[2] in ("Très fort", "Plutôt fort") and num(row[5]):
            series.setdefault(cycle, {}).setdefault(terr, 0.0)
            series[cycle][terr] += float(row[5])
    cycles = sorted(series)
    appart = {"cycles": cycles,
              "mtl": [round(series[c].get("mtl"), 1) for c in cycles],
              "qc": [round(series[c].get("qc"), 1) for c in cycles]}

    # satisfaction vie sociale par RTS — deux cycles (2014-2015, 2020-2021)
    ws = wb["satis-vie-sociale CIUSSS"]
    cats = ["Très satisfaisante", "Plutôt satisfaisante",
            "Plutôt insatisfaisante", "Très insatisfaisante"]
    tables, cur, current_terr = [], None, None
    for row in ws.iter_rows(values_only=True):
        c0 = str(row[0]) if row[0] else ""
        if c0.startswith("Répartition"):
            cur = {}
            tables.append(cur)
            current_terr = None
        if cur is None:
            continue
        m = re.match(r"^(06\d)\s*-", c0)
        if m:
            current_terr = m.group(1)
        elif c0.startswith("Total de la région"):
            current_terr = "mtl"
        if current_terr and row[1] in cats and num(row[5]):
            cur.setdefault(current_terr, {})[cats.index(str(row[1]).replace("\n", " "))] = round(float(row[5]), 1)
    satisf = {"cats": cats,
              "c2014": {t: [v[i] for i in range(4)] for t, v in tables[0].items()} if tables else {},
              "c2020": {t: [v[i] for i in range(4)] for t, v in tables[1].items()} if len(tables) > 1 else {}}

    # degré de solitude (EQSP 2020-2021)
    ws = wb["degré-solitude"]
    solitude = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] == "06 Montréal" and num(row[3]):
            solitude["mtl"] = {"moy": round(float(row[3]), 2),
                               "ic": [round(float(row[6]), 2), round(float(row[7]), 2)]}
        elif row[0] == "Ensemble du Québec" and num(row[3]):
            solitude["qc"] = {"moy": round(float(row[3]), 2),
                              "ic": [round(float(row[6]), 2), round(float(row[7]), 2)]}

    print(f"capital : appartenance {len(cycles)} cycles | satisfaction {len(tables)} cycles | solitude {list(solitude)}")
    return {"appartenance": appart, "satisfaction": satisf, "solitude": solitude}


# --------- Données saisies à la main (sources publiques, voir méta) ----------
# Insécurité alimentaire des ménages, Québec, ECR 2018-2023 (ISQ, mars 2026) ;
# Montréal 2023 : 22 % (même publication).
ALIM_DATA = {
    "meta": {"source": "Institut de la statistique du Québec, « L'insécurité alimentaire au "
                       "Québec entre 2018 et 2023 » (Enquête canadienne sur le revenu), mars 2026",
             "url": "https://statistique.quebec.ca/fr/produit/publication/insecurite-alimentaire-2018-2023-faits-saillants"},
    "years": [2018, 2019, 2020, 2021, 2022, 2023],
    "qc": [14.0, 11.6, 13.1, 13.8, 16.4, 19.0],
    "mtl2023": 22.0,
    "laval2023": 23.0,
}

# Taux de participation, élection municipale du 7 novembre 2021, par
# arrondissement (mairie d'arrondissement ; Ville-Marie : moyenne pondérée des
# trois districts). Source : Élections Montréal, résultats officiels.
PARTICIPATION_2021 = {
    "meta": {"source": "Élections Montréal, élection générale du 7 novembre 2021 — résultats officiels",
             "overall": 38.3, "annee": 2021},
    "geo": {  # slugs de la carte Ville de Montréal (arrondissements seulement)
        "ahuntsic": 42.3, "anjou": 40.3, "cdnndg": 34.2, "ilebizard": 49.0,
        "lachine": 39.2, "lasalle": 30.9, "mhm": 41.6, "mtlnord": 33.5,
        "outremont": 55.9, "pfdsrox": 30.5, "plateau": 44.7, "rdp": 39.0,
        "rosemont": 47.9, "stlaurent": 29.0, "stleonard": 32.7, "sudouest": 35.1,
        "verdun": 42.8, "villemarie": 34.0, "villeray": 37.0,
    },
}


# ------------------------------------------------------------- agrégation ----
def new_defavo():
    return {"pop": 0, "nad": 0, "mat": [0.0] * 5, "soc": [0.0] * 5}


def add_defavo(agg, slug, rec):
    a = agg.setdefault(slug, new_defavo())
    a["pop"] += rec["pop"]
    a["nad"] += 1
    a["mat"][rec["qm"] - 1] += rec["pop"]
    a["soc"][rec["qs"] - 1] += rec["pop"]


def pct(x, pop):
    return round(100 * x / pop, 1)


def q45(a, dim):
    return None if not a or a["pop"] == 0 else pct(a[dim][3] + a[dim][4], a["pop"])


def main():
    iemv = load_iemv()
    tq_of = tq_assigner()
    vdm_slug = vdm_name_to_slug()
    rls2021 = load_rls_2021()

    # correspondances AD 2021 -> territoire, pour chaque découpage
    unmatched_vdm = {slugify(v["nom"]) for v in iemv.values()
                     if slugify(v["nom"]) not in vdm_slug}
    if unmatched_vdm:
        print("WARN noms IEMV sans slug Ville :", unmatched_vdm)
    geo_of_2021 = {
        "tq": {ad: tq_of(v["pt"]) for ad, v in iemv.items()},
        "vdm": {ad: vdm_slug.get(slugify(v["nom"])) for ad, v in iemv.items()},
        "sq": {ad: ("rls-" + rls2021[ad]) if ad in rls2021 else None for ad in iemv},
    }

    # arbre des points 2021 pour rattacher les AD 2011/2016 (découpage vdm)
    ads21 = list(iemv)
    tree = STRtree([iemv[ad]["pt"] for ad in ads21])

    def vdm_of_old(pt):
        return geo_of_2021["vdm"][ads21[tree.nearest(pt)]]

    # ---------------- defavo : 3 années × 3 découpages ----------------
    years_data = {}
    for year in (2011, 2016, 2021):
        inspq = load_inspq(year)
        if inspq is None:
            print(f"{year} : table INSPQ absente — année sautée")
            continue
        if year == 2021:
            pts = {ad: iemv[ad]["pt"] for ad in iemv}
        else:
            pts = points_statcan(year)
            if pts is None:
                print(f"{year} : géométrie absente — année sautée")
                continue
        agg = {"tq": {}, "vdm": {}, "sq": {}}
        miss = 0
        for ad, rec in inspq.items():
            add_defavo(agg["sq"], "rls-" + rec["rls"], rec)
            pt = pts.get(ad)
            if pt is None:
                miss += 1
                continue
            tq = geo_of_2021["tq"].get(ad) if year == 2021 else tq_of(pt)
            if tq:
                add_defavo(agg["tq"], tq, rec)
            vd = geo_of_2021["vdm"].get(ad) if year == 2021 else vdm_of_old(pt)
            if vd:
                add_defavo(agg["vdm"], vd, rec)
        years_data[year] = agg
        print(f"defavo {year} : {len(inspq)} AD (sans géométrie : {miss}) | "
              f"tq {len(agg['tq'])} | vdm {len(agg['vdm'])} | sq {len(agg['sq'])} territoires")

    if 2021 not in years_data:
        sys.exit("ERREUR : l'année 2021 est requise.")
    years = sorted(years_data)

    defavo_geo = {}
    for geo in ("tq", "vdm", "sq"):
        out = {}
        for slug, a in sorted(years_data[2021][geo].items()):
            if a["pop"] == 0:
                continue
            rec = {
                "pop": a["pop"], "nad": a["nad"],
                "mat": [pct(x, a["pop"]) for x in a["mat"]],
                "soc": [pct(x, a["pop"]) for x in a["soc"]],
            }
            if len(years) > 1:
                rec["trend"] = {
                    "years": years,
                    "mat": [q45(years_data[y][geo].get(slug), "mat") for y in years],
                    "soc": [q45(years_data[y][geo].get(slug), "soc") for y in years],
                }
            out[slug] = rec
        defavo_geo[geo] = out

    # ---------------- iemv : version 2026, 3 découpages ----------------
    dim_keys = [k for k, _f, _l in IEMV_DIMS]
    iemv_geo = {}
    for geo in ("tq", "vdm", "sq"):
        agg = {}
        for ad, v in iemv.items():
            slug = geo_of_2021[geo].get(ad)
            if slug is None or v["score"] is None or v["pop"] == 0:
                continue
            a = agg.setdefault(slug, {"pop": 0, "nad": 0, "dist": [0.0] * 7,
                                      "niv": [0.0, 0.0, 0.0],
                                      "dims": {k: [0.0, 0.0] for k in dim_keys}})
            a["pop"] += v["pop"]
            a["nad"] += 1
            a["dist"][v["score"]] += v["pop"]
            cls = iemv_class(None, v["score"])
            a["niv"][{"nv": 0, "vnp": 1, "pv": 2}[cls]] += v["pop"]
            for k in dim_keys:
                if v["acp"].get(k) == 1:
                    a["dims"][k][0] += v["pop"]              # vulnérable (dimension)
                    if v["score"] >= 4:
                        a["dims"][k][1] += v["pop"]          # ... et zone prioritaire
        out = {}
        for slug, a in sorted(agg.items()):
            if a["pop"] == 0:
                continue
            out[slug] = {
                "pop": a["pop"], "nad": a["nad"],
                "dist": [pct(x, a["pop"]) for x in a["dist"]],
                "p4": pct(sum(a["dist"][4:]), a["pop"]),
                "niv": [pct(x, a["pop"]) for x in a["niv"]],
                "dims": {k: [pct(x, a["pop"]) for x in a["dims"][k]] for k in dim_keys},
            }
        iemv_geo[geo] = out
        tp = sum(a["pop"] for a in out.values())
        t4 = sum(a["pop"] * a["p4"] / 100 for a in out.values())
        print(f"iemv {geo} : {len(out)} territoires | pop {tp:,} | ≥4 vulnérabilités {100*t4/tp:.1f} %")

    # totaux île pour l'IEMV (comparaison dans le panneau)
    iemv_tot = {"pop": 0, "niv": [0.0, 0.0, 0.0], "dims": {k: [0.0, 0.0] for k in dim_keys}}
    for v in iemv.values():
        if v["score"] is None or v["pop"] == 0:
            continue
        iemv_tot["pop"] += v["pop"]
        cls = iemv_class(None, v["score"])
        iemv_tot["niv"][{"nv": 0, "vnp": 1, "pv": 2}[cls]] += v["pop"]
        for k in dim_keys:
            if v["acp"].get(k) == 1:
                iemv_tot["dims"][k][0] += v["pop"]
                if v["score"] >= 4:
                    iemv_tot["dims"][k][1] += v["pop"]
    iemv_overall = {
        "niv": [pct(x, iemv_tot["pop"]) for x in iemv_tot["niv"]],
        "dims": {k: [pct(x, iemv_tot["pop"]) for x in iemv_tot["dims"][k]] for k in dim_keys},
    }

    # ---------------- iemv : mosaïque d'AD dissoute (fichier géométrie) --------
    iemv_paths = build_iemv_geo(iemv)

    # ---------------- mpc : taux MPC par AD (champ MPC21 du GeoJSON IEMV) --------
    # moyenne des taux d'AD pondérée par la population de l'AD
    mpc_data = {"geo": {}}
    mpc_tot = [0.0, 0.0]
    for geo in ("tq", "vdm", "sq"):
        m = {}
        for ad, v in iemv.items():
            if v["mpc"] is None or v["pop"] == 0:
                continue
            if geo == "tq":  # totaux île : une seule fois
                mpc_tot[0] += v["mpc"] * v["pop"]
                mpc_tot[1] += v["pop"]
            slug = geo_of_2021[geo].get(ad)
            if slug is None:
                continue
            r = m.setdefault(slug, [0.0, 0.0, 0])
            r[0] += v["mpc"] * v["pop"]
            r[1] += v["pop"]
            r[2] += 1
        mpc_data["geo"][geo] = {s: {"v": round(a / b, 1), "pop": int(b), "nad": n}
                                for s, (a, b, n) in sorted(m.items()) if b}
    mpc_data["meta"] = {
        "source": "Statistique Canada, Recensement 2021 ; champ MPC21 du jeu de données "
                  "IEMV (Ville de Montréal)",
        "overall": round(mpc_tot[0] / mpc_tot[1], 1) if mpc_tot[1] else None,
        "annee": 2021,
    }
    print(f"mpc : île {mpc_data['meta']['overall']} % | " +
          " | ".join(f"{g} {len(mpc_data['geo'][g])} terr." for g in ("tq", "vdm", "sq")))

    # ---------------- logement : taux d'effort >= 30 % (profil du recensement) ---
    # cache produit par tools/extract_census2021.py ; comptes de ménages :
    # CID 1465 = base (ménages propriétaires et locataires, revenu > 0),
    # CID 1467 = 30 % ou plus du revenu consacré aux frais de logement.
    logement_data = None
    cache = IND / "census2021_mtl.csv"
    if cache.exists():
        num = lambda s: None if s in ("", "..", "x", "F", "...") else float(s)
        byad = {}
        for row in csv.DictReader(open(cache, encoding="utf-8")):
            byad.setdefault(row["DAUID"], {})[row["CID"]] = num(row["C1"])
        logement_data = {"geo": {}}
        log_tot = [0.0, 0.0]
        for geo in ("tq", "vdm", "sq"):
            lg = {}
            for ad, vals in byad.items():
                base, c30 = vals.get("1465"), vals.get("1467")
                if not base or c30 is None:
                    continue
                if geo == "tq":
                    log_tot[0] += c30
                    log_tot[1] += base
                slug = geo_of_2021[geo].get(ad)
                if slug is None:
                    continue
                r = lg.setdefault(slug, [0.0, 0.0, 0])
                r[0] += c30
                r[1] += base
                r[2] += 1
            logement_data["geo"][geo] = {s: {"v": round(100 * a / b, 1), "men": int(b), "nad": n}
                                         for s, (a, b, n) in sorted(lg.items()) if b}
        logement_data["meta"] = {
            "source": "Statistique Canada, Recensement 2021 (98-401-X2021006), "
                      "profil des aires de diffusion",
            "overall": round(100 * log_tot[0] / log_tot[1], 1) if log_tot[1] else None,
            "annee": 2021,
        }
        print(f"logement 30 %+ : île {logement_data['meta']['overall']} % | " +
              " | ".join(f"{g} {len(logement_data['geo'][g])} terr." for g in ("tq", "vdm", "sq")))
    else:
        print("census2021_mtl.csv absent — indicateur logement sauté "
              "(lancer tools/extract_census2021.py)")

    # contrôle defavo : part régionale Q4-Q5 ≈ 40 %
    for y in years:
        agg = years_data[y]["sq"]
        tp = sum(a["pop"] for a in agg.values())
        tm = sum(a["mat"][3] + a["mat"][4] for a in agg.values())
        print(f"contrôle defavo {y} (sq) : pop {tp:,} | Q4-Q5 mat {100*tm/tp:.1f} % (attendu ≈ 40 %)")

    data = {}
    if mpc_data:
        data["mpc"] = mpc_data
    if logement_data:
        data["logement"] = logement_data
    data |= {
        "defavo": {
            "meta": {
                "source": "INSPQ, Indice de défavorisation matérielle et sociale (2011, 2016, 2021) ; "
                          "Statistique Canada, recensements",
                "reference": "Quintiles régionaux (RSS 06 — Montréal)",
                "annees": years,
            },
            "geo": defavo_geo,
        },
        "iemv": {
            "meta": {
                "source": "Ville de Montréal, Indice d'équité des milieux de vie, version 2026",
                "annee": 2026,
                "dims": {k: lbl for k, _f, lbl in IEMV_DIMS},
                "overall": iemv_overall,
            },
            "geo": iemv_geo,
        },
        "alim": ALIM_DATA,
        "participation": PARTICIPATION_2021,
    }
    capital = load_capital()
    if capital:
        data["capital"] = capital
        data["capital"]["meta"] = {
            "appartenance": "Statistique Canada, ESCC 2007-2008 à 2019-2020 ; Infocentre de santé publique (INSPQ). "
                            "Population de 12 ans et plus déclarant un sentiment d'appartenance très ou plutôt fort à sa communauté locale.",
            "satisfaction": "EQSP 2014-2015 et 2020-2021 ; Infocentre de santé publique (INSPQ). Population de 15 ans et plus. "
                            "Les deux cycles ne sont pas directement comparables.",
            "solitude": "EQSP 2020-2021 ; Infocentre de santé publique (INSPQ). Degré moyen de solitude "
                        "(population de 15 ans et plus ; un score plus élevé indique un degré de solitude plus grand).",
        }
    js = ("/* Indicateurs — DONNÉES (généré par tools/build_indicateurs.py, ne pas éditer).\n"
          "   INDIC_DATA.defavo : IDMS par territoire (mat/soc = % de population par quintile\n"
          "   régional 2021 ; trend = % en quintiles 4-5 par recensement).\n"
          "   INDIC_DATA.iemv : IEMV 2026 (dist = % par nombre de vulnérabilités 0-6 ;\n"
          "   p4 = % en zone vulnérable et prioritaire ; niv = % [non vuln., vuln. non\n"
          "   prioritaire, vuln. et prioritaire] ; dims = % [vulnérable, vulnérable et\n"
          "   prioritaire] par dimension).\n"
          "   INDIC_DATA.alim / capital / participation : voir les méta de chaque bloc.\n"
          "   Découpages : tq (tables de quartier), vdm (arrondissements et villes liées),\n"
          "   sq (réseaux locaux de services). */\n"
          "const INDIC_DATA = " +
          json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    OUT.write_text(js, encoding="utf-8")
    print(f"Écrit : {OUT}")

    jsg = ("/* Indicateurs — GÉOMÉTRIE IEMV (généré par tools/build_indicateurs.py, ne pas\n"
           "   éditer). Aires de diffusion 2021 dissoutes par option (ens = ensemble,\n"
           "   cult/prox/secu/envi/eco/soci = dimensions) et par classe :\n"
           "   pv = vulnérable et prioritaire, vnp = vulnérable non prioritaire,\n"
           "   nv = non vulnérable. Espace SVG 1000x875 (même projection que les cartes). */\n"
           "const IEMV_GEO = " +
           json.dumps(iemv_paths, ensure_ascii=False, separators=(",", ":")) + ";\n")
    OUT_GEO.write_text(jsg, encoding="utf-8")
    print(f"Écrit : {OUT_GEO} ({OUT_GEO.stat().st_size/1024:.0f} ko)")


if __name__ == "__main__":
    main()
